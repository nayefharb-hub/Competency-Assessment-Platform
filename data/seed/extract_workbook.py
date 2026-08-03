#!/usr/bin/env python3
"""T0 — extract the KIB ICB4 workbook into verified seed data + a verification report."""
import openpyxl, json, re, sys, os

SRC = "/root/.claude/uploads/2c8d4b22-3d44-5fa4-bc77-355111d631c6/b41f4e37-KIB_PM_Competency_Framework_ICB4_v1_1.xlsx"
OUT = "/home/user/Competency-Assessment-Platform/data/seed"
os.makedirs(OUT, exist_ok=True)

wb = openpyxl.load_workbook(SRC, data_only=True)
report = []
def log(s):
    report.append(s); print(s)

# ---------- SCALE ----------
ws = wb["Scale"]
scale = []
for row in ws.iter_rows(min_row=4, max_row=9, values_only=True):
    if row[0] is None: continue
    scale.append({"level": int(row[0]), "label": str(row[1]).strip(),
                  "knowledge": str(row[2]).strip(), "application": str(row[3]).strip()})
log(f"SCALE: {len(scale)} levels -> {[s['label'] for s in scale]}")

# ---------- BENCHMARKS ----------
ws = wb["Benchmarks"]
benchmarks = []
for row in ws.iter_rows(min_row=5, max_row=33, values_only=True):
    if row[0] is None or row[1] is None: continue
    def lv(v):
        if v is None or str(v).strip() == "": return None
        s = str(v).strip()
        return s if s == "N/R" else int(float(s))
    benchmarks.append({"no": int(row[0]), "apm_competence": str(row[1]).strip(),
                       "entry": lv(row[2]), "intermediate": lv(row[3]),
                       "advanced": lv(row[4]), "master": lv(row[5])})
log(f"BENCHMARKS: {len(benchmarks)} APM competences (expect 29)")

# ---------- FRAMEWORK ----------
ws = wb["Framework"]
AREAS = {"PERSPECTIVE": "Perspective", "PEOPLE": "People", "PRACTICE": "Practice"}
ce_re = re.compile(r"^(\d+\.\d+\.\d+)\s+(.+?)\s+—\s+(\d+)\s+controls?\s*$")
ctrl_re = re.compile(r"^\d+\.\d+\.\d+\.\d+$")

area = None; ce = None
areas, ces, controls = [], [], []
for row in ws.iter_rows(min_row=4, values_only=True):
    a = row[0]
    if a is None: continue
    s = str(a).strip()
    if s.upper() in AREAS:
        area = AREAS[s.upper()]
        areas.append({"code": area.lower(), "name": area})
        continue
    m = ce_re.match(re.sub(r"\s+", " ", s))
    if m:
        ce = {"code": m.group(1), "name": m.group(2).strip(), "area": area,
              "declared_controls": int(m.group(3))}
        ces.append(ce); continue
    if ctrl_re.match(s):
        def cell(i):
            v = row[i]
            if v is None: return None
            t = str(v).strip()
            # the workbook writes a literal "None" string for empty reason/note cells
            return None if t == "" or t == "None" else t
        active_raw = row[8]
        active = True if active_raw is None else (str(active_raw).strip().upper() in ("TRUE", "YES", "1"))
        tnum = row[13]   # N = numeric target level (0-5)
        controls.append({
            "code": s, "ce_code": ce["code"], "area": area,
            "indicator": cell(1), "description": cell(2),
            "active": active,
            "priority": cell(9), "reason": cell(10), "kib_note": cell(11),
            "apm_competence": cell(12),
            "target_level": None if tnum is None or str(tnum).strip()=="" else int(float(tnum)),
            "target_label": cell(14),   # O = label for that level
            "target_source": cell(15),
        })

log(f"AREAS: {len(areas)} -> {[a['name'] for a in areas]}")
log(f"COMPETENCE ELEMENTS: {len(ces)} (expect 28)")
log(f"CONTROLS: {len(controls)} (expect 133)")
active_n = sum(1 for c in controls if c["active"])
log(f"  active={active_n} (expect 132) · inactive={len(controls)-active_n} (expect 1)")

# ---------- MEASURES ----------
ws = wb["Measures"]
measures = []
for row in ws.iter_rows(min_row=4, values_only=True):
    if row[0] is None: continue
    measures.append({"control_code": str(row[0]).strip(),
                     "ce_name": str(row[1]).strip() if row[1] else None,
                     "no": int(row[2]) if row[2] is not None else None,
                     "text": str(row[3]).strip() if row[3] else None})
log(f"MEASURES: {len(measures)} rows")

# ---------- RESULTS (CE targets + rollup rules) ----------
ws = wb["Results"]
ce_targets = []
for row in ws.iter_rows(min_row=4, max_row=31, values_only=True):
    if row[1] is None: continue
    ce_targets.append({"area": str(row[0]).strip(), "ce_code": str(row[1]).strip(),
                       "ce_name": str(row[2]).strip(),
                       "active_controls": int(row[3]) if row[3] is not None else None,
                       "target": int(row[4]) if row[4] is not None else None})
log(f"CE TARGETS (Results sheet): {len(ce_targets)} (expect 28)")

# ---------- VERIFICATION ----------
log("\n--- VERIFICATION ---")
ok = True
def check(label, got, want):
    global ok
    good = got == want
    ok = ok and good
    log(f"  [{'OK ' if good else 'FAIL'}] {label}: got {got}, expect {want}")
    return good

check("controls total", len(controls), 133)
check("controls active", active_n, 132)
check("competence elements", len(ces), 28)
check("areas", len(areas), 3)
check("CE targets rows", len(ce_targets), 28)

# per-area KCI counts from the handover brief: Perspective 24, People 49, Practice 60
for a, want in (("Perspective", 24), ("People", 49), ("Practice", 60)):
    check(f"{a} controls", sum(1 for c in controls if c["area"] == a), want)
# per-area CE counts: 5/10/13
for a, want in (("Perspective", 5), ("People", 10), ("Practice", 13)):
    check(f"{a} competence elements", sum(1 for c in ces if c["area"] == a), want)

# Per-CE header counts declare ACTIVE controls (verified: 4.3.2 shows "6 controls"
# for 7 rows because 4.3.2.6 is inactive). Check against the active count.
mismatch = []
for ce_ in ces:
    n = sum(1 for c in controls if c["ce_code"] == ce_["code"] and c["active"])
    if n != ce_["declared_controls"]:
        mismatch.append((ce_["code"], ce_["declared_controls"], n))
log(f"  [{'OK ' if not mismatch else 'FAIL'}] per-CE declared vs ACTIVE control counts: {len(mismatch)} mismatches")
for m in mismatch: log(f"       {m[0]}: header says {m[1]}, active {m[2]}")

# Results-sheet active-control counts must agree with the Framework sheet
rs_mismatch = []
for t in ce_targets:
    n = sum(1 for c in controls if c["ce_code"] == t["ce_code"] and c["active"])
    if t["active_controls"] is not None and t["active_controls"] != n:
        rs_mismatch.append((t["ce_code"], t["active_controls"], n))
log(f"  [{'OK ' if not rs_mismatch else 'FAIL'}] Results vs Framework active counts: {len(rs_mismatch)} mismatches")
for m in rs_mismatch: log(f"       {m[0]}: Results says {m[1]}, Framework has {m[2]}")

# CE codes align between Framework and Results
fw_codes = {c["code"] for c in ces}; rs_codes = {c["ce_code"] for c in ce_targets}
log(f"  [{'OK ' if fw_codes==rs_codes else 'FAIL'}] CE codes match Framework vs Results")
if fw_codes != rs_codes:
    log(f"       only in Framework: {sorted(fw_codes-rs_codes)}")
    log(f"       only in Results:   {sorted(rs_codes-fw_codes)}")

# measures reference real controls
ctrl_codes = {c["code"] for c in controls}
orphan = {m["control_code"] for m in measures} - ctrl_codes
log(f"  [{'OK ' if not orphan else 'FAIL'}] measures reference known controls ({len(orphan)} orphans)")

# every active control has a target + priority
no_target = [c["code"] for c in controls if c["active"] and c["target_level"] is None]
no_prio   = [c["code"] for c in controls if c["active"] and not c["priority"]]
log(f"  [{'OK ' if not no_target else 'FAIL'}] active controls with a target level ({len(no_target)} missing)")
log(f"  [{'OK ' if not no_prio else 'FAIL'}] active controls with a priority ({len(no_prio)} missing)")
if no_target[:5]: log(f"       e.g. {no_target[:5]}")

# priority distribution (brief: High 112, Medium 17, Low 3)
from collections import Counter
pc = Counter(c["priority"] for c in controls if c["active"])
log(f"  priority distribution (active): {dict(pc)}  [brief: High 112, Medium 17, Low 3]")
# target source split (brief: APM 101, derived 31)
ts = Counter(c["target_source"] for c in controls if c["active"])
log(f"  target source (active): {dict(ts)}  [brief: APM-published 101, derived 31]")
inactive = [c["code"] for c in controls if not c["active"]]
log(f"  inactive control(s): {inactive}  [brief: 4.3.2.6]")

# ---------- WRITE ----------
bundle = {"source_file": os.path.basename(SRC), "framework": "IPMA ICB4 v4.0.1",
          "scale": {"name": "APM Competence Framework 3rd Edition", "axis": "Application", "levels": scale},
          "areas": areas, "competence_elements": ces, "controls": controls,
          "measures": measures, "benchmarks": benchmarks, "ce_targets": ce_targets}
with open(f"{OUT}/icb4-framework.json", "w", encoding="utf8") as f:
    json.dump(bundle, f, ensure_ascii=False, indent=2)
log(f"\nWROTE {OUT}/icb4-framework.json")

import csv
with open(f"{OUT}/controls.csv", "w", newline="", encoding="utf8") as f:
    w = csv.DictWriter(f, fieldnames=list(controls[0].keys())); w.writeheader(); w.writerows(controls)
with open(f"{OUT}/measures.csv", "w", newline="", encoding="utf8") as f:
    w = csv.DictWriter(f, fieldnames=list(measures[0].keys())); w.writeheader(); w.writerows(measures)
log(f"WROTE {OUT}/controls.csv, {OUT}/measures.csv")

with open(f"{OUT}/EXTRACTION-REPORT.txt", "w", encoding="utf8") as f:
    f.write("\n".join(report) + "\n")
print("\nRESULT:", "ALL CHECKS PASSED" if ok and not mismatch and not orphan else "SEE FAILURES ABOVE")
